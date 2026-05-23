import pandas as pd
import requests
import ipaddress
import os
import json
import math
import xml.etree.ElementTree as ET
from datetime import datetime
from geopy.distance import geodesic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
import threading
import time

# Caching for IP Geolocation to prevent redundant API calls
IP_CACHE = {}

# ip-api.com free tier: 45 requests/minute
_IPAPI_SEMAPHORE = threading.Semaphore(5)
_IPAPI_LOCK = threading.Lock()
_IPAPI_LAST_CALL = [0.0]
_IPAPI_MIN_INTERVAL = 1.4

class PMLogEngine:
    def __init__(self, gmaps_api_key=None, abuse_api_key=None, proxy_settings=None, 
                 segment_xml_path=None, enable_abuseipdb=True, enable_geolocation=True, 
                 enable_impossible_travel=True, enable_routing_api=True, segments_only_mode=False,
                 internal_prefix_len=24, external_prefix_len=24, 
                 subnet_switch_suspicious_mins=30, subnet_switch_review_mins=120,
                 summarize_switch_details=True):
        self.gmaps_api_key = gmaps_api_key
        self.abuse_api_key = abuse_api_key
        self.proxy_settings = proxy_settings
        self.internal_prefix_len = internal_prefix_len
        self.external_prefix_len = external_prefix_len
        self.subnet_switch_suspicious_mins = subnet_switch_suspicious_mins
        self.subnet_switch_review_mins = subnet_switch_review_mins
        self.summarize_switch_details = summarize_switch_details
        
        # Operational Toggles
        self.segments_only_mode = segments_only_mode
        if self.segments_only_mode:
            self.enable_abuseipdb = False
            self.enable_geolocation = False
            self.enable_impossible_travel = False
            self.enable_routing_api = False
        else:
            self.enable_abuseipdb = enable_abuseipdb
            self.enable_geolocation = enable_geolocation
            self.enable_impossible_travel = enable_impossible_travel
            self.enable_routing_api = enable_routing_api

        # Geolocation is required for impossible travel
        if not self.enable_geolocation and self.enable_impossible_travel:
            self.enable_impossible_travel = False

        self.proxies = self._build_proxy()
        
        # XML Segment Map
        self.segment_xml_path = segment_xml_path
        self.segment_map = []
        if self.segment_xml_path and os.path.exists(self.segment_xml_path):
            self.segment_map = self._load_segment_xml(self.segment_xml_path)

    def _build_proxy(self):
        if not self.proxy_settings or not self.proxy_settings.get('enabled'):
            return None
        p = self.proxy_settings
        auth = ""
        if p.get('user') and p.get('pass'):
            auth = f"{p['user']}:{p['pass']}@"
        proxy_url = f"http://{auth}{p.get('host', '')}:{p.get('port', '')}"
        return {"http": proxy_url, "https": proxy_url}

    def _load_segment_xml(self, xml_path):
        """Parses Segmentnoenv.xml and returns a flat list of segment entries with location hierarchy."""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            results = []

            def parse_groups(element, parent_location=None, current_group_name=None, depth=0):
                name = element.get('NAME')
                
                # A "Location" is usually a top-level group under GTBANK (depth 2 or 3)
                # Depth 0: Segments, Depth 1: GTB-ENV, Depth 2: GTBANK, Depth 3: Locations (Lagos, North, etc.)
                is_location = False
                if depth == 3:
                    is_location = True
                
                loc = name if is_location else parent_location
                group_name = name if name else current_group_name
                
                # Check for RANGES children
                for child in element:
                    if child.tag == 'RANGES':
                        range_str = child.get('RANGE')
                        if not range_str: continue
                        
                        try:
                            is_hyphen = '-' in range_str
                            start_ip = None
                            end_ip = None
                            net_obj = None
                            
                            if is_hyphen:
                                parts = range_str.split('-')
                                start_ip = ipaddress.ip_address(parts[0].strip())
                                end_ip = ipaddress.ip_address(parts[1].strip())
                            else:
                                if '/' not in range_str:
                                    net_obj = ipaddress.ip_network(range_str.strip() + '/32', strict=False)
                                else:
                                    net_obj = ipaddress.ip_network(range_str.strip(), strict=False)
                                    
                            results.append({
                                'location': loc or "GTBANK",
                                'segment': group_name,
                                'is_hyphen': is_hyphen,
                                'start': start_ip,
                                'end': end_ip,
                                'net': net_obj
                            })
                        except:
                            continue 
                            
                    elif child.tag == 'GROUP':
                        parse_groups(child, loc, group_name, depth + 1)

            parse_groups(root)
            return results
        except Exception as e:
            print(f"Error loading segment XML: {e}")
            return []

    def get_segment_info(self, ip_str):
        """Identifies which location and named segment an IP belongs to."""
        if not self.segment_map:
            return None, None
        try:
            ip_obj = ipaddress.ip_address(ip_str.strip())
            if not ip_obj.is_private:
                return None, None
                
            for entry in self.segment_map:
                if entry['is_hyphen']:
                    if entry['start'] <= ip_obj <= entry['end']:
                        return entry['location'], entry['segment']
                else:
                    if ip_obj in entry['net']:
                        return entry['location'], entry['segment']
            return "GTBANK", "Unknown Segment"
        except:
            return None, None

    def get_segment_name(self, ip_str):
        """Returns the hierarchical name for the IP."""
        loc, seg = self.get_segment_info(ip_str)
        if not loc: return None
        if loc == seg: return loc
        return f"{loc} - {seg}"

    def get_location_context(self, ip_str, segment_info_str):
        """Identifies the functional context of a login (e.g., ATM, Server, Staff Desk)."""
        try:
            ip_obj = ipaddress.ip_address(ip_str.strip())
            
            if not ip_obj.is_private:
                return "External Network"
                
            last_octet = int(str(ip_obj).split('.')[-1])
            
            # Heuristics for context
            if last_octet == 225:
                return "Self-Service Kiosk"
            if 215 <= last_octet <= 224:
                return "PINPAD Terminal"
            if 213 <= last_octet <= 214:
                return "POS Device"
            
            if segment_info_str:
                info_upper = segment_info_str.upper()
                if "ATM" in info_upper: return "ATM Terminal"
                if "SERVER" in info_upper or "DATACENTER" in info_upper: return "Server/Admin"
                if "VPN" in info_upper: return "Remote VPN"
                if "WIFI" in info_upper: return "Wireless Client"
                if "VOICE" in info_upper: return "VoIP Device"
                if "DEV" in info_upper: return "Developer Workstation"
            
            return "Standard Workstation"
        except:
            return "Unknown Context"

    def is_external(self, ip_str):
        try:
            ip = ipaddress.ip_address(ip_str.strip())
            return not ip.is_private
        except:
            return False

    def is_same_network(self, ip1: str, ip2: str) -> bool:
        """Compares two IPs using XML segment names if available, else prefix lengths."""
        try:
            addr1 = ipaddress.ip_address(ip1.strip())
            addr2 = ipaddress.ip_address(ip2.strip())
        except:
            return True # Safe default

        # One private, one public -> Always different
        if addr1.is_private != addr2.is_private:
            return False

        # If both private and XML available, use segment names
        if addr1.is_private and self.segment_map:
            name1 = self.get_segment_name(ip1)
            name2 = self.get_segment_name(ip2)
            if name1 and name2:
                return name1 == name2
            # Fall back to prefix comparison if name not found in XML

        prefix = self.internal_prefix_len if addr1.is_private else self.external_prefix_len
        try:
            net1 = ipaddress.ip_network(f"{ip1.strip()}/{prefix}", strict=False)
            net2 = ipaddress.ip_network(f"{ip2.strip()}/{prefix}", strict=False)
            return net1 == net2
        except:
            return True
        try:
            net1 = ipaddress.ip_network(f"{ip1.strip()}/{prefix}", strict=False)
            net2 = ipaddress.ip_network(f"{ip2.strip()}/{prefix}", strict=False)
            return net1 == net2
        except:
            return True

    def classify_segment(self, seg):
        try:
            return "External" if self.is_external(seg + ".1") else "Internal"
        except:
            return "Unknown"

    def get_location(self, ip):
        if not self.enable_geolocation:
            return None

        if ip in IP_CACHE:
            return IP_CACHE[ip]

        if not self.is_external(ip):
            return None

        url = f"http://ip-api.com/json/{ip}?fields=status,message,countryCode,city,lat,lon"
        max_retries = 3

        with _IPAPI_SEMAPHORE:
            for attempt in range(max_retries):
                with _IPAPI_LOCK:
                    now = time.time()
                    wait = _IPAPI_MIN_INTERVAL - (now - _IPAPI_LAST_CALL[0])
                    if wait > 0:
                        time.sleep(wait)
                    _IPAPI_LAST_CALL[0] = time.time()

                try:
                    resp = requests.get(url, proxies=self.proxies, timeout=10)
                    if resp.status_code == 429:
                        retry_after = int(resp.headers.get('Retry-After', 60))
                        print(f"[ip-api] Rate limited on {ip}. Waiting {retry_after}s (attempt {attempt+1}/{max_retries})")
                        time.sleep(retry_after)
                        continue
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get('status') == 'success':
                            IP_CACHE[ip] = data
                            return data
                    return None
                except Exception as e:
                    print(f"[ip-api] Error fetching {ip}: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
            return None

    def check_abuseipdb(self, ip):
        """Check an IP against AbuseIPDB. Returns a dict with abuse score and category info."""
        if not self.enable_abuseipdb or not self.abuse_api_key:
            return None
        if not self.is_external(ip):
            return None
        try:
            url = "https://api.abuseipdb.com/api/v2/check"
            headers = {"Key": self.abuse_api_key, "Accept": "application/json"}
            params = {"ipAddress": ip, "maxAgeInDays": 90, "verbose": True}
            resp = requests.get(url, headers=headers, params=params, proxies=self.proxies, timeout=10)
            if resp.status_code == 200:
                data = resp.json().get("data", {})
                return {
                    "IP": ip,
                    "Abuse Confidence Score": data.get("abuseConfidenceScore", 0),
                    "Total Reports": data.get("totalReports", 0),
                    "Last Reported": data.get("lastReportedAt") or "N/A",
                    "Country": data.get("countryCode", ""),
                    "ISP": data.get("isp", ""),
                    "Usage Type": data.get("usageType", ""),
                    "Is Public": data.get("isPublic", False),
                    "Is Tor": data.get("isTor", False),
                    "Classification": "Malicious" if data.get("abuseConfidenceScore", 0) >= 75 else (
                        "Suspicious" if data.get("abuseConfidenceScore", 0) >= 25 else "Clean")
                }
            return None
        except Exception:
            return None

    def get_gmaps_travel_time_hours(self, lat1, lon1, lat2, lon2):
        if not self.enable_routing_api or not self.gmaps_api_key:
            return None
        try:
            url = "https://maps.googleapis.com/maps/api/distancematrix/json"
            params = {
                "origins": f"{lat1},{lon1}",
                "destinations": f"{lat2},{lon2}",
                "key": self.gmaps_api_key
            }
            resp = requests.get(url, params=params, proxies=self.proxies, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("rows") and data["rows"][0].get("elements"):
                    element = data["rows"][0]["elements"][0]
                    if element.get("status") == "OK":
                        # Return travel time in hours
                        return element["duration"]["value"] / 3600.0
            return None
        except Exception:
            return None

    def get_osrm_travel_time_hours(self, lat1, lon1, lat2, lon2):
        """Use free OSRM public routing API (no key required) to get driving time."""
        if not self.enable_routing_api:
            return None
        try:
            # OSRM expects lon,lat order
            url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"
            params = {"overview": "false", "steps": "false"}
            resp = requests.get(url, params=params, proxies=self.proxies, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("code") == "Ok" and data.get("routes"):
                    # Duration is in seconds
                    return data["routes"][0]["duration"] / 3600.0
            return None
        except Exception:
            return None

    def get_haversine_travel_time_hours(self, lat1, lon1, lat2, lon2):
        # Calculate straight-line distance in km
        distance_km = geodesic((lat1, lon1), (lat2, lon2)).kilometers
        # Assuming typical commercial flight speed of 900 km/h
        # Also, minimum realistic time is distance / 900, plus ~2 hours for airport overhead.
        # But for impossible travel strictly, distance / 900 is the absolute physical limit.
        return distance_km / 900.0

    def _get_isp_key(self, ip, abuse_by_ip):
        """Returns (isp, country) tuple for external IP clustering if data is available."""
        if not abuse_by_ip or ip not in abuse_by_ip:
            return None
        res = abuse_by_ip[ip]
        isp = str(res.get('ISP', '')).strip().lower()
        country = str(res.get('Country', '')).strip().upper()
        if not isp or not country: return None
        return (isp, country)

    def _extract_unique_ips(self, file_path):
        """Fast pass to extract all unique external IPs for pre-fetching context."""
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == '.csv':
                df = pd.read_csv(file_path, header=1)
            else:
                df = pd.read_excel(file_path, header=1)
            
            unique_ips = set()
            for _, row in df.iterrows():
                ip_str = str(row.get('LOG IP LIST', ''))
                ips = [ip.strip() for ip in ip_str.split(',') if ip.strip()]
                for ip in ips:
                    if self.is_external(ip):
                        unique_ips.add(ip)
            return list(unique_ips)
        except:
            return []

    def _fetch_abuse_batch(self, ips):
        """Fetch reputation for a list of IPs in parallel."""
        if not self.abuse_api_key or not ips:
            return {}
        
        abuse_by_ip = {}
        # Limit to 10 parallel workers to avoid overloading or rate limits
        with ThreadPoolExecutor(max_workers=10) as executor:
            results = list(executor.map(self.check_abuseipdb, ips))
            for res in results:
                if res and 'IP' in res:
                    abuse_by_ip[res['IP']] = res
        return abuse_by_ip

    def _format_mail_lines(self, parsed_events: list, abuse_by_ip: dict = None, min_switch_pair: tuple = (None, None)) -> str:
        if not parsed_events:
            return ""
            
        # Group by cluster, picking the best timestamp
        cluster_map = {} # cluster_idx -> event
        
        # Priority 1: Events from the min switch pair
        if min_switch_pair[0]:
            ev1, ev2 = min_switch_pair
            c1 = ev1.get('info', {}).get('cluster_idx') if 'info' in ev1 else ev1['ip']
            c2 = ev2.get('info', {}).get('cluster_idx') if 'info' in ev2 else ev2['ip']
            cluster_map[c1] = ev1
            cluster_map[c2] = ev2
            
        # Priority 2: Latest event for any cluster not yet mapped
        for ev in sorted(parsed_events, key=lambda x: x['time']):
            cid = ev.get('info', {}).get('cluster_idx') if 'info' in ev else ev['ip']
            if cid not in cluster_map or ev['time'] > cluster_map[cid]['time']:
                # If cid is in cluster_map but from min_switch_pair, don't overwrite it
                # unless this event is also from the pair (which is handled by Priority 1).
                # Actually, if it's already in cluster_map, it's either the min_switch one or we already picked a later one.
                if cid not in cluster_map:
                    cluster_map[cid] = ev
                elif min_switch_pair[0] and (ev is min_switch_pair[0] or ev is min_switch_pair[1]):
                    cluster_map[cid] = ev
                elif not (min_switch_pair[0] and (cluster_map[cid] is min_switch_pair[0] or cluster_map[cid] is min_switch_pair[1])):
                     # Only overwrite if the existing one is NOT part of the min switch pair
                     if ev['time'] > cluster_map[cid]['time']:
                         cluster_map[cid] = ev
            
        lines = []
        # Sort clusters by the time we picked for them
        sorted_clusters = sorted(cluster_map.values(), key=lambda x: x['time'])
        
        for ev in sorted_clusters:
            ip = ev['ip']
            name = self.get_segment_name(ip)
            if not name and abuse_by_ip and ip in abuse_by_ip:
                isp_key = self._get_isp_key(ip, abuse_by_ip)
                if isp_key:
                    name = f"{isp_key[0].title()} ({isp_key[1]})"

            if not name:
                name = ip
            else:
                prefixes = ["LAGOS - ", "NORTH - ", "WEST - ", "MIDWEST AND SOUTH - ", "OFFSITE - "]
                for p in prefixes:
                    if name.upper().startswith(p):
                        name = name[len(p):].strip()
                        break
            
            # Format: 9:24:05 AM (12-hour, stripped leading zero)
            formatted_time = ev['time'].strftime('%I:%M:%S %p').lstrip('0')
            lines.append(f"{name}: {ip} --- {formatted_time}")
                
        return "\n".join(lines)

    def parse_logs(self, file_path, abuse_by_ip=None):
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == '.csv':
                df = pd.read_csv(file_path, header=1)
            else:
                df = pd.read_excel(file_path, header=1)
        except Exception as e:
            raise ValueError(f"Failed to read file: {str(e)}")

        mismatch_data = []
        impossible_travel_data = []
        unique_external_ips = set()

        for _, row in df.iterrows():
            username = str(row.get('USR USERNAME', 'Unknown'))
            cnt = row.get('CNT', 0)
            ip_str = str(row.get('LOG IP LIST', ''))
            date_str = str(row.get('LOG INIT DATE LIST', ''))
            
            ips = [ip.strip() for ip in ip_str.split(',') if ip.strip()]
            dates_raw = [d.strip() for d in date_str.split(',') if d.strip()]

            # Clean and parse dates
            parsed_events = []
            for i, ip in enumerate(ips):
                if self.is_external(ip):
                    unique_external_ips.add(ip)

                if i < len(dates_raw):
                    try:
                        # Assuming format like "2026-03-07 04:27:32"
                        dt = datetime.strptime(dates_raw[i], "%Y-%m-%d %H:%M:%S")
                        parsed_events.append({"ip": ip, "time": dt})
                    except ValueError:
                        pass # Ignore invalid dates

            # Segmentation Check (Mismatches) - ISP-Aware Clustering
            network_clusters = []
            ip_to_cluster_info = {}

            for ip in ips:
                found_cluster_idx = -1
                is_ext = self.is_external(ip)
                my_key = self._get_isp_key(ip, abuse_by_ip) if is_ext else None
                
                for idx, cluster in enumerate(network_clusters):
                    lead = cluster[0]
                    lead_is_ext = self.is_external(lead)
                    
                    if is_ext and lead_is_ext:
                        lead_key = self._get_isp_key(lead, abuse_by_ip)
                        if my_key and my_key == lead_key:
                            found_cluster_idx = idx
                            break
                        elif not my_key and self.is_same_network(ip, lead):
                            found_cluster_idx = idx
                            break
                    elif not is_ext and not lead_is_ext:
                        if self.is_same_network(ip, lead):
                            found_cluster_idx = idx
                            break
                
                if found_cluster_idx == -1:
                    found_cluster_idx = len(network_clusters)
                    network_clusters.append([ip])
                else:
                    network_clusters[found_cluster_idx].append(ip)
                
                # Store cluster info for this IP
                addr = ipaddress.ip_address(ip)
                prefix = self.internal_prefix_len if addr.is_private else self.external_prefix_len
                net_str = str(ipaddress.ip_network(f"{ip}/{prefix}", strict=False))
                
                label = net_str
                seg_name = None
                location = None
                if is_ext and my_key:
                    label = f"{my_key[0].title()} ({my_key[1]})"
                elif not is_ext:
                    location, segment = self.get_segment_info(ip)
                    seg_name = self.get_segment_name(ip)
                    if seg_name:
                        label = f"{net_str} ({seg_name})"
                
                ip_to_cluster_info[ip] = {
                    "cluster_idx": found_cluster_idx,
                    "label": label,
                    "network": net_str,
                    "location": location or ("External" if is_ext else "Internal"),
                    "segment_name": seg_name
                }
            
            if len(network_clusters) > 1:
                networks_found = [c[0] for c in network_clusters] # Just leads
                classifications = {self.classify_segment('.'.join(c[0].split('.')[:3])) for c in network_clusters}
                
                # VPN Check: One internal cluster + one external cluster
                is_vpn_case = False
                if len(network_clusters) == 2:
                    types = [self.is_external(c[0]) for c in network_clusters]
                    if types.count(True) == 1 and types.count(False) == 1:
                        is_vpn_case = True
                
                if not is_vpn_case:
                    # Time Context Analysis
                    switch_events = []
                    min_delta_seconds = None
                    
                    # user_events uses parsed_events which already handles positional alignment (skip if no date)
                    user_events = []
                    for ev in parsed_events:
                        ip = ev['ip']
                        if ip in ip_to_cluster_info:
                            user_events.append({
                                'ip': ip,
                                'time': ev['time'],
                                'info': ip_to_cluster_info[ip]
                            })
                    
                    # Handle edge case: User has only one timestamp but multiple IPs (from different clusters)
                    # The requirement says delta_seconds = 0 if same time but multiple locations.
                    # Our parsed_events only includes IPs that HAVE a date.
                    # If multiple IPs have the SAME date, we'll get deltas of 0.
                    
                    # Sort by time
                    user_events.sort(key=lambda x: x['time'])
                    
                    min_switch_pair = (None, None)
                    for i in range(len(user_events) - 1):
                        ev1 = user_events[i]
                        ev2 = user_events[i+1]
                        
                        if ev1['info']['cluster_idx'] != ev2['info']['cluster_idx']:
                            delta = int((ev2['time'] - ev1['time']).total_seconds())
                            switch_events.append({
                                'from_ip': ev1['ip'],
                                'from_cluster': ev1['info']['label'],
                                'to_ip': ev2['ip'],
                                'to_cluster': ev2['info']['label'],
                                'delta': delta
                            })
                            if min_delta_seconds is None or delta < min_delta_seconds:
                                min_delta_seconds = delta
                                min_switch_pair = (ev1, ev2)

                    # Formatting
                    def format_delta(secs):
                        if secs is None: return "N/A"
                        h = secs // 3600
                        m = (secs % 3600) // 60
                        s = secs % 60
                        return f"{h}h {m:02d}m {s:02d}s"

                    min_switch_time_str = format_delta(min_delta_seconds)
                    
                    feasibility = "N/A"
                    if min_delta_seconds is not None:
                        min_mins = min_delta_seconds / 60.0
                        if min_mins < self.subnet_switch_suspicious_mins:
                            feasibility = "Suspicious"
                        elif min_mins < self.subnet_switch_review_mins:
                            feasibility = "Review"
                        else:
                            feasibility = "Plausible"
                    elif len(network_clusters) > 1:
                        feasibility = "Unknown"

                    if self.summarize_switch_details:
                        transition_summary = {}
                        for sw in switch_events:
                            path_key = f"{sw['from_cluster']} → {sw['to_cluster']}"
                            if path_key not in transition_summary:
                                transition_summary[path_key] = {'count': 0, 'min_delta': float('inf')}
                            
                            transition_summary[path_key]['count'] += 1
                            if sw['delta'] < transition_summary[path_key]['min_delta']:
                                transition_summary[path_key]['min_delta'] = sw['delta']
                                
                        detail_lines = []
                        for path, stats in transition_summary.items():
                            detail_lines.append(f"{path} ({stats['count']}x) | fastest: {format_delta(stats['min_delta'])}")
                    else:
                        detail_lines = []
                        for sw in switch_events:
                            detail_lines.append(f"{sw['from_ip']} ({sw['from_cluster']}) → {sw['to_ip']} ({sw['to_cluster']}) | delta: {format_delta(sw['delta'])}")
                        
                    switch_detail = "\n".join(detail_lines)

                    isp_groups = []
                    for cluster in network_clusters:
                        if self.is_external(cluster[0]):
                            key = self._get_isp_key(cluster[0], abuse_by_ip)
                            if key:
                                isp_groups.append(f"{key[0].title()} ({key[1]})")

                    mismatch_entry = {
                        'Username': username,
                        'Total Logins': cnt,
                        'Location(s)': ", ".join(sorted(set(ip_to_cluster_info[ip]['location'] for ip in ips))),
                        'Segments Found': ", ".join(sorted(set(ip_to_cluster_info[ip]['label'] for ip in ips))),
                        'Min Segment Switch Time': min_switch_time_str,
                        'Feasibility': feasibility,
                        'Switch Detail': switch_detail,
                    }
                    if self.enable_abuseipdb:
                        mismatch_entry['ISP Groups'] = " | ".join(sorted(set(isp_groups))) if isp_groups else ""
                    
                    # Mail Lines: Exactly N lines, one per cluster, capturing the min switch if possible
                    mismatch_entry['Mail Lines'] = self._format_mail_lines(user_events, abuse_by_ip, min_switch_pair)
                    mismatch_entry['Full IP List'] = ", ".join(ips)
                    log_date_short = min(e['time'] for e in user_events).strftime('%d %b %Y') if user_events else "Unknown Date"
                    mismatch_entry['Email Subject'] = f"Multiple Login Activity Observed on Processmaker — {username} ({log_date_short})"
                    mismatch_data.append(mismatch_entry)

            # Impossible Travel Check
            if self.enable_impossible_travel and len(parsed_events) > 1:
                # Sort events by time
                parsed_events.sort(key=lambda x: x['time'])
                
                for i in range(len(parsed_events) - 1):
                    ev1 = parsed_events[i]
                    ev2 = parsed_events[i+1]
                    
                    if ev1['ip'] == ev2['ip']:
                        continue # Same IP, no travel

                    loc1 = self.get_location(ev1['ip'])
                    loc2 = self.get_location(ev2['ip'])

                    if loc1 and loc2 and (loc1['lat'] != loc2['lat'] or loc1['lon'] != loc2['lon']):
                        time_diff_hours = (ev2['time'] - ev1['time']).total_seconds() / 3600.0
                        
                        # Only flag if time difference is extremely short (e.g. less than 24 hours)
                        # and they changed physical locations
                        if time_diff_hours < 24:
                            travel_time_hours = None
                            calc_method = ""
                            
                            # 1. Try Google Maps
                            if self.gmaps_api_key:
                                travel_time_hours = self.get_gmaps_travel_time_hours(
                                    loc1['lat'], loc1['lon'], loc2['lat'], loc2['lon']
                                )
                                if travel_time_hours is not None:
                                    calc_method = "Google Maps (Driving)"

                            # 2. Try OSRM (free, no key needed) if GMaps not used/failed
                            if travel_time_hours is None:
                                travel_time_hours = self.get_osrm_travel_time_hours(
                                    loc1['lat'], loc1['lon'], loc2['lat'], loc2['lon']
                                )
                                if travel_time_hours is not None:
                                    calc_method = "OSRM (Free Driving Route)"

                            # 3. Fallback to Haversine Flight Time
                            if travel_time_hours is None:
                                travel_time_hours = self.get_haversine_travel_time_hours(
                                    loc1['lat'], loc1['lon'], loc2['lat'], loc2['lon']
                                )
                                calc_method = "Haversine Fallback (900km/h Minimum Flight)"

                            # Flag if actual time took less than physical minimum time
                            if time_diff_hours < travel_time_hours:
                                distance_km = round(geodesic((loc1['lat'], loc1['lon']), (loc2['lat'], loc2['lon'])).kilometers, 2)
                                impossible_travel_data.append({
                                    'Username': username,
                                    'IP 1': ev1['ip'],
                                    'Location 1': f"{loc1.get('city', '')}, {loc1.get('countryCode', '')}",
                                    'Time 1': ev1['time'].strftime("%Y-%m-%d %H:%M:%S"),
                                    'IP 2': ev2['ip'],
                                    'Location 2': f"{loc2.get('city', '')}, {loc2.get('countryCode', '')}",
                                    'Time 2': ev2['time'].strftime("%Y-%m-%d %H:%M:%S"),
                                    'Physical Distance (km)': distance_km,
                                    'Time Passed (hrs)': round(time_diff_hours, 2),
                                    'Min Possible Travel Time (hrs)': round(travel_time_hours, 2),
                                    'Calculation Method': calc_method
                                })

        return df, mismatch_data, impossible_travel_data, list(unique_external_ips)

    def generate_report(self, file_path, output_path=None, version=2, abuse_by_ip=None):
        """Routing function for versioned report generation."""
        if version == 2:
            return self._generate_report_v2(file_path, output_path, abuse_by_ip=abuse_by_ip)
        return self._generate_report_v1(file_path, output_path, abuse_by_ip=abuse_by_ip)

    def _generate_report_v1(self, file_path, output_path=None, abuse_by_ip=None):
        # Pre-fetch ISP data for roaming filter if not provided
        if abuse_by_ip is None:
            unique_ips = self._extract_unique_ips(file_path)
            abuse_by_ip = self._fetch_abuse_batch(unique_ips)

        df_orig, mismatches, impossible_travel, external_ips = self.parse_logs(file_path, abuse_by_ip=abuse_by_ip)

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(os.path.dirname(file_path), f"PM_Analysis_{timestamp}.xlsx")

        # AbuseIPDB check on all unique external IPs
        abuse_results = []
        if self.enable_abuseipdb and self.abuse_api_key:
            for ip in sorted(external_ips):
                result = self.check_abuseipdb(ip)
                if result:
                    abuse_results.append(result)

        # Clean columns with trailing commas before writing
        clean_cols = ['LOG IP LIST', 'LOG INIT DATE LIST', 'LOG END DATE LIST', 'LOG CLIENT HOSTNAME LIST']
        for col in clean_cols:
            if col in df_orig.columns:
                df_orig[col] = df_orig[col].apply(lambda x: ', '.join([v.strip() for v in str(x).split(',') if v.strip()]))

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_orig.to_excel(writer, sheet_name='Original Data', index=False)
            
            df_mismatch = pd.DataFrame(mismatches)
            if not df_mismatch.empty:
                df_mismatch.to_excel(writer, sheet_name='Subnet Anomalies', index=False)
            else:
                pd.DataFrame({"Message": ["No subnet anomalies found"]}).to_excel(writer, sheet_name='Subnet Anomalies', index=False)
                
            df_impossible = pd.DataFrame(impossible_travel)
            if not df_impossible.empty:
                df_impossible.to_excel(writer, sheet_name='Impossible Travel Alerts', index=False)
            else:
                pd.DataFrame({"Message": ["No impossible travel detected"]}).to_excel(writer, sheet_name='Impossible Travel Alerts', index=False)

            pd.DataFrame({'UNIQUE EXTERNAL IPS': sorted(external_ips)}).to_excel(writer, sheet_name='Unique External IPs', index=False)

            # AbuseIPDB sheet
            df_abuse = pd.DataFrame(abuse_results)
            if not df_abuse.empty:
                df_abuse.to_excel(writer, sheet_name='AbuseIPDB Reputation', index=False)
            elif self.abuse_api_key:
                pd.DataFrame({"Message": ["All checked IPs returned clean or no data."]}).to_excel(writer, sheet_name='AbuseIPDB Reputation', index=False)

        return output_path

    def _generate_report_v2(self, file_path, output_path=None, abuse_by_ip=None):
        # 1. Pre-fetch ISP data for roaming filter and enrichment if not provided
        abuse_results = []
        if abuse_by_ip is None:
            unique_ips = self._extract_unique_ips(file_path)
            abuse_by_ip = self._fetch_abuse_batch(unique_ips)
            abuse_results = list(abuse_by_ip.values())
        else:
            # Use provided dict
            abuse_results = list(abuse_by_ip.values())

        df_orig, mismatches, impossible_travel, external_ips = self.parse_logs(file_path, abuse_by_ip=abuse_by_ip)

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(os.path.dirname(file_path), f"PM_Analysis_{timestamp}.xlsx")

        # 2. Enrichment Mappings (Already populated above)

        # Map IPs to users for traceability
        ip_to_users = {}
        for _, row in df_orig.iterrows():
            uname = str(row.get('USR USERNAME', 'Unknown'))
            ips = [ip.strip() for ip in str(row.get('LOG IP LIST', '')).split(',') if ip.strip()]
            for ip in ips:
                if self.is_external(ip):
                    if ip not in ip_to_users: ip_to_users[ip] = set()
                    ip_to_users[ip].add(uname)

        # Clean columns with trailing commas before writing
        clean_cols = ['LOG IP LIST', 'LOG INIT DATE LIST', 'LOG END DATE LIST', 'LOG CLIENT HOSTNAME LIST']
        for col in clean_cols:
            if col in df_orig.columns:
                df_orig[col] = df_orig[col].apply(lambda x: ', '.join([v.strip() for v in str(x).split(',') if v.strip()]))

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. Prepare All Alerts
            all_alerts = []
            filtered_mismatches = []
            
            for m in mismatches:
                ips_involved = [ip.strip() for ip in m['Full IP List'].split(',') if ip.strip()]
                segments_found_str = str(m.get('Segments Found', ''))
                
                # Gather unique contexts across all IPs in this mismatch
                unique_contexts = set()
                for ip in ips_involved:
                    _, seg_name = self.get_segment_info(ip)
                    ctx = self.get_location_context(ip, str(seg_name))
                    seg_upper = str(seg_name).upper()
                    if "VPC" in seg_upper or "VIRTUAL PRIVATE" in seg_upper or "VIRTUAL PC" in seg_upper:
                        ctx = "VPC"
                    unique_contexts.add(ctx)
                
                # Rule 1: Exclude Server/Admin or Developer Workstation
                if "Server/Admin" in unique_contexts or "Developer Workstation" in unique_contexts:
                    continue
                
                # Rule 2: Exclude VPC if it's just Workstation and VPC (<= 2 network groups)
                if "VPC" in unique_contexts:
                    allowed_workstations = {"Standard Workstation", "Wireless Client", "Remote VPN", "Unknown Context", "VPC"}
                    if unique_contexts.issubset(allowed_workstations) and m.get('Unique Network Groups', 0) <= 2:
                        continue
                
                # Determine primary context for report display
                primary_context = self.get_location_context(ips_involved[0], segments_found_str)
                seg_upper_str = segments_found_str.upper()
                if "VPC" in seg_upper_str or "VIRTUAL PRIVATE" in seg_upper_str or "VIRTUAL PC" in seg_upper_str:
                    primary_context = "VPC"
                    
                filtered_mismatches.append(m)

                feasibility = m.get('Feasibility', 'N/A')
                risk = 'Low'
                if feasibility == 'Suspicious': risk = 'High'
                elif feasibility == 'Review': risk = 'Medium'
                elif feasibility == 'Plausible': risk = 'Low'
                else:
                    seg_types = m.get('Segment Types', '')
                    is_mixed = 'Internal' in seg_types and 'External' in seg_types
                    risk = 'Medium' if is_mixed else 'Low'
                
                countries = []
                isps = []
                for ip in ips_involved:
                    if ip in abuse_by_ip:
                        res = abuse_by_ip[ip]
                        if res.get('Country'): countries.append(res['Country'])
                        if res.get('ISP'): isps.append(res['ISP'])
                
                alert = {
                    "Username": m['Username'],
                    "Logins": m['Total Logins'],
                    "Alert Type": "Subnet Anomaly",
                    "Feasibility": feasibility,
                    "Location": m.get('Location(s)', 'Internal'),
                }
                if self.enable_abuseipdb:
                    alert["Countries"] = ", ".join(sorted(set(countries)))
                    alert["ISP(s)"] = ", ".join(sorted(set(isps)))
                alert.update({
                    "Min Switch": m['Min Segment Switch Time'],
                    "Subnet Detail": segments_found_str,
                    "IPs Involved": m['Full IP List'],
                    "Mail Lines": m.get('Mail Lines', ''),
                    "Email Subject": m.get('Email Subject', '')
                })
                all_alerts.append(alert)
            
            mismatches = filtered_mismatches
            
            for t in impossible_travel:
                alert = {
                    "Username": t['Username'],
                    "Logins": 2,
                    "Alert Type": "Impossible Travel",
                    "Feasibility": "Critical",
                    "Location": f"{t['Location 1']} -> {t['Location 2']}",
                }
                if self.enable_abuseipdb:
                    alert["Countries"] = f"{t['Location 1']}, {t['Location 2']}"
                    alert["ISP(s)"] = ""
                alert.update({
                    "Min Switch": f"{round(t['Time Passed (hrs)'], 2)} hrs",
                    "Subnet Detail": f"{t['IP 1']} to {t['IP 2']}",
                    "IPs Involved": f"{t['IP 1']}, {t['IP 2']}",
                    "Mail Lines": "",
                    "Email Subject": ""
                })
                all_alerts.append(alert)
            
            feasibility_rank = {"Critical": 1, "Suspicious": 2, "Review": 3, "Plausible": 4, "Unknown": 5, "N/A": 6}
            all_alerts.sort(key=lambda x: feasibility_rank.get(x['Feasibility'], 99))

            # 2. Write Sheets
            pd.DataFrame().to_excel(writer, sheet_name='Summary') # Placeholder
            if all_alerts:
                pd.DataFrame(all_alerts).to_excel(writer, sheet_name='Anomalies', index=False)
            if mismatches:
                pd.DataFrame(mismatches).to_excel(writer, sheet_name='Subnet Anomalies', index=False)
            if impossible_travel:
                pd.DataFrame(impossible_travel).to_excel(writer, sheet_name='Impossible Travel', index=False)
            if abuse_results:
                pd.DataFrame(abuse_results).to_excel(writer, sheet_name='IP Intelligence', index=False)
            df_orig.to_excel(writer, sheet_name='Original Data', index=False)

        # 3. Apply Professional Formatting
        from openpyxl.formatting.rule import CellIsRule
        wb = openpyxl.load_workbook(output_path)
        
        # Summary Formatting
        ws_sum = wb['Summary']
        summary_items = [
            ["PROCESSMAKER LOG ANALYSIS REPORT", ""],
            ["Generated At", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Source File", os.path.basename(file_path)],
            ["", ""],
            ["EXECUTIVE SUMMARY", ""],
            ["Total Events Analyzed", len(df_orig)],
            ["Unique Users", df_orig['USR USERNAME'].nunique() if 'USR USERNAME' in df_orig.columns else 0],
            ["Total External IPs Detected", len(external_ips)],
            ["Subnet Anomalies Found", len(mismatches)],
            ["Impossible Travel Alerts", len(impossible_travel)],
            ["Malicious IPs Found", len([r for r in abuse_results if r.get('Classification') == 'Malicious'])],
            ["", ""],
            ["ANALYSIS MODE", "Full" if not self.segments_only_mode else "Offline"],
            ["XML Segment Map", os.path.basename(self.segment_xml_path) if self.segment_xml_path else "Not used"],
            ["", ""]
        ]
        
        if all_alerts:
            summary_items.append(["TOP ANOMALOUS LOCATIONS", "Alert Count"])
            loc_counts = {}
            for a in all_alerts:
                primary_loc = a['Location'].split(' - ')[0].split(' -> ')[0]
                loc_counts[primary_loc] = loc_counts.get(primary_loc, 0) + 1
            for loc, count in sorted(loc_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                summary_items.append([loc, count])

        for r_idx, item in enumerate(summary_items, 1):
            ws_sum.cell(row=r_idx, column=1, value=item[0]).font = Font(bold=True)
            ws_sum.cell(row=r_idx, column=2, value=item[1])
        ws_sum.column_dimensions['A'].width = 35
        ws_sum.column_dimensions['B'].width = 60

        # Formatting All Sheets
        sheets_to_process = [
            ('Anomalies', 'AnomaliesTable', 'TableStyleMedium2'),
            ('Subnet Anomalies', 'SubnetAnomalies', 'TableStyleMedium2'),
            ('Impossible Travel', 'ImpossibleTravel', 'TableStyleMedium2'),
            ('IP Intelligence', 'IPIntelligence', 'TableStyleMedium2'),
            ('Original Data', None, None) # No table for original
        ]

        for sname, tname, style in sheets_to_process:
            if sname in wb.sheetnames:
                ws = wb[sname]
                self._bold_header_row(ws)
                self._auto_fit_columns(ws)
                ws.freeze_panes = "A2"
                if tname:
                    self._apply_table(ws, tname, style)
                
                # Restore full row color highlighting
                fill_map = {}
                if sname == 'Anomalies':
                    headers = [cell.value for cell in ws[1]]
                    try:
                        feas_idx = headers.index("Feasibility")
                    except ValueError:
                        feas_idx = -1
                    try:
                        mail_lines_idx = headers.index("Mail Lines")
                    except ValueError:
                        mail_lines_idx = -1
                    try:
                        email_subj_idx = headers.index("Email Subject")
                    except ValueError:
                        email_subj_idx = -1
                        
                    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
                        if feas_idx != -1:
                            feas = str(row[feas_idx].value)
                            if feas in ['Suspicious', 'Critical']: fill_map[i] = 'FFC7CE'
                            elif feas == 'Review': fill_map[i] = 'FFEB9C'
                            
                        if mail_lines_idx != -1:
                            row[mail_lines_idx].alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                            row[mail_lines_idx].font = Font(size=10)
                            ws.row_dimensions[i].height = None
                            
                        if email_subj_idx != -1:
                            row[email_subj_idx].font = Font(size=10)
                            
                elif sname == 'Subnet Anomalies':
                    headers = [cell.value for cell in ws[1]]
                    try:
                        feas_idx = headers.index("Feasibility")
                    except ValueError:
                        feas_idx = -1
                    try:
                        detail_idx = headers.index("Switch Detail")
                    except ValueError:
                        detail_idx = -1
                    try:
                        mail_lines_idx = headers.index("Mail Lines")
                    except ValueError:
                        mail_lines_idx = -1
                    try:
                        email_subj_idx = headers.index("Email Subject")
                    except ValueError:
                        email_subj_idx = -1

                    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
                        if feas_idx != -1:
                            feas = str(row[feas_idx].value)
                            if feas == 'Suspicious': fill_map[i] = 'FFC7CE'
                            elif feas == 'Review': fill_map[i] = 'FFEB9C'
                            elif feas == 'Plausible': fill_map[i] = 'C6EFCE'
                        
                        if detail_idx != -1:
                            row[detail_idx].alignment = Alignment(wrap_text=True, vertical='top')
                            ws.row_dimensions[i].height = None
                            
                        if mail_lines_idx != -1:
                            row[mail_lines_idx].alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                            row[mail_lines_idx].font = Font(size=10)
                            ws.row_dimensions[i].height = None
                            
                        if email_subj_idx != -1:
                            row[email_subj_idx].font = Font(size=10)
                            
                elif sname == 'Impossible Travel':
                    for i in range(2, ws.max_row + 1): fill_map[i] = 'FFC7CE'
                    
                elif sname == 'IP Intelligence':
                    headers = [cell.value for cell in ws[1]]
                    try:
                        ut_idx = headers.index("Usage Type")
                    except ValueError:
                        ut_idx = -1
                    try:
                        cls_idx = headers.index("Classification")
                    except ValueError:
                        cls_idx = -1
                    try:
                        tor_idx = headers.index("Is Tor")
                    except ValueError:
                        tor_idx = -1

                    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
                        ut = str(row[ut_idx].value) if ut_idx != -1 else ""
                        cls = str(row[cls_idx].value) if cls_idx != -1 else ""
                        tor = row[tor_idx].value if tor_idx != -1 else False
                        
                        if cls == 'Malicious' or tor is True: fill_map[i] = 'FFC7CE'
                        elif "Data Center" in ut: fill_map[i] = 'FFEB9C'
                
                if fill_map:
                    self._apply_row_fills(ws, fill_map)

        wb.save(output_path)
        return output_path



    def _auto_fit_columns(self, worksheet):
        """Iterate over worksheet columns and adjust width based on longest content."""
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get column letter
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 4, 60) # Cap at 60
            worksheet.column_dimensions[column].width = adjusted_width

    def _apply_table(self, worksheet, table_name, style='TableStyleMedium2'):
        """Convert a range of data into an Excel Table."""
        if worksheet.max_row < 2:
            return # Don't create empty tables
        
        # Table names must be unique and alphanumeric
        safe_name = "".join(filter(str.isalnum, table_name))
        ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        tab = Table(displayName=safe_name, ref=ref)
        
        style_info = TableStyleInfo(name=style, showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style_info
        worksheet.add_table(tab)

    def _bold_header_row(self, worksheet):
        """Applies bold font to the first row."""
        bold_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = bold_font

    def _apply_row_fills(self, worksheet, fill_map, start_row=2):
        """Apply PatternFill to rows based on a row_index -> hex mapping."""
        for row_idx, color_hex in fill_map.items():
            if not color_hex: continue
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            for cell in worksheet[row_idx]:
                cell.fill = fill
